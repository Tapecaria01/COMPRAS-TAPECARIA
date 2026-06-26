import streamlit as st
import pandas as pd
import pdfplumber
import re
import math
import traceback
import plotly.express as px
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Portal Compras - Tapeçaria", layout="wide")

# ==========================================
# --- ESTILIZAÇÃO CSS (DESIGN PREMIUM) ---
# ==========================================
st.markdown("""
    <style>
    .stApp { background-color: #0E1117 !important; }
    div.stButton > button:first-child {
        background-color: #004A8F !important;
        color: white !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 8px 24px !important;
        font-weight: 600 !important;
        transition: 0.3s !important;
    }
    div.stButton > button:first-child:hover { 
        background-color: #003366 !important; 
        color: white !important; 
    }
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# --- TELA DE SENHA (BLOQUEIO CENTRALIZADO) ---
# ==========================================
SENHA_ACESSO = "Tape2026"

if "liberado" not in st.session_state:
    st.session_state.liberado = False

if not st.session_state.liberado:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br><br><br>", unsafe_allow_html=True)
        sc1, sc2, sc3 = st.columns([1.5, 1, 1.5])
        with sc2:
            try: 
                st.image("logo.png", use_container_width=True)
            except: 
                st.markdown("<h1 style='text-align: center; color: white;'>🏢</h1>", unsafe_allow_html=True)
            
        st.markdown("<h2 style='text-align: center; color: #FFFFFF;'>Acesso Restrito</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: #4DA8DA;'>Insira a senha de sistema para aceder à inteligência de compras.</p>", unsafe_allow_html=True)
        
        senha = st.text_input("Senha", type="password")
        if st.button("Entrar no Portal", use_container_width=True):
            if senha == SENHA_ACESSO:
                st.session_state.liberado = True
                st.rerun()
            else:
                st.error("Senha incorreta. Tente novamente.")
    st.stop()

# ==========================================
# --- VARIÁVEIS DE SESSÃO ---
# ==========================================
if "df_regras" not in st.session_state:
    st.session_state.df_regras = pd.DataFrame([
        {"FORNECEDOR": "CORTTEX", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "TEX COMPANY", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "CIPATEX", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "KARSTEN", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "ETRURIA", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "TELLAIO", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "OBER", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "TEXTIL J. SERRANO", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "CKS", "MULTIPLO": 50, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "AGRO QUIMICA", "MULTIPLO": 45, "TOLERANCIA": 20, "PALAVRA_CHAVE": ""},
        {"FORNECEDOR": "ROMPLAS", "MULTIPLO": 30, "TOLERANCIA": 15, "PALAVRA_CHAVE": "URUGUA"},
        {"FORNECEDOR": "ROMA DUBLADOS", "MULTIPLO": 10, "TOLERANCIA": 5, "PALAVRA_CHAVE": ""}
    ])

# ==========================================
# --- FUNÇÕES DE APOIO ---
# ==========================================
def limpar_v(valor):
    if not valor: 
        return 0.0
    s = str(valor).strip().replace('.', '').replace(',', '.')
    try: 
        numero_limpo = re.sub(r'[^\d.]', '', s)
        return float(numero_limpo)
    except: 
        return 0.0

# --- MELHORIA 1: CACHE INTELIGENTE DE PDF ---
# Guarda a extração pesada na memória para o recálculo ser instantâneo
@st.cache_data(show_spinner=False)
def extrair_dados_pdf_web(pdf_file):
    dados = []
    nome_filial = pdf_file.name.replace(".pdf", "").upper()
    meses_encontrados = []
    fornecedor_atual = "DESCONHECIDO"
    
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if not texto: 
                    continue
                    
                if len(meses_encontrados) < 4:
                    padrao_mes = r'\b(?:jan|fev|feb|mar|abr|apr|mai|may|jun|jul|ago|aug|set|sep|out|oct|nov|dez|dec)/\d{2,4}\b'
                    encontrados = re.findall(padrao_mes, texto.lower())
                    for m in encontrados:
                        m_upper = m.upper()
                        if m_upper not in meses_encontrados: 
                            meses_encontrados.append(m_upper)
                            
                for linha in texto.split('\n'):
                    l = linha.strip()
                    if "SEGMENTO" in l.upper():
                        match = re.search(r'SEGMENTO\s*:\s*(.*)', l, re.IGNORECASE)
                        if match: 
                            fornecedor_atual = match.group(1).strip()
                    else:
                        if "YORK" in fornecedor_atual.upper() and l.startswith("***"):
                            l = re.sub(r'^\*\*\*\s*', '', l)
                            
                        if re.match(r'^\d{3,6}\s', l):
                            partes = l.split()
                            try:
                                item_dict = {
                                    'CODIGO': partes[0], 
                                    'DESCRICAO': " ".join(partes[1:-11]), 
                                    'EMB.': partes[-11],
                                    'MES_1': limpar_v(partes[-10]), 
                                    'MES_2': limpar_v(partes[-9]), 
                                    'MES_3': limpar_v(partes[-8]),
                                    'MES_4': limpar_v(partes[-7]), 
                                    'MEDIA_SISTEMA': limpar_v(partes[-6]), 
                                    'ESTOQUE': limpar_v(partes[-5]),
                                    'RESERVA': limpar_v(partes[-4]), 
                                    'COMPRADA': limpar_v(partes[-3]), 
                                    'MESES_ESTOQUE': limpar_v(partes[-1]), 
                                    'FILIAL_NOME': nome_filial, 
                                    'FORNECEDOR': fornecedor_atual
                                }
                                dados.append(item_dict)
                            except: 
                                continue
        return pd.DataFrame(dados), meses_encontrados
    except Exception as e: 
        return pd.DataFrame(), []

# --- INTERFACE WEB (BARRA LATERAL) ---
with st.sidebar:
    try: 
        st.image("logo.png", use_container_width=True)
    except: 
        pass
        
    st.markdown("---")
    st.header("📂 Nova Compra")
    uploaded_files = st.file_uploader("Selecione os 4 PDFs das Unidades", type="pdf", accept_multiple_files=True)
    st.markdown("---")
    
    with st.expander("⚙️ Configurações Avançadas"):
        meta = st.number_input("Meta de estoque (meses)", min_value=1, value=2)
        meses_parado = st.number_input("Considerar estoque parado após (meses)", min_value=1, value=3, step=1)
        fator_pico = st.number_input("Sensibilidade de Pico (x vezes a média)", min_value=1.5, value=2.5, step=0.5)
        nome_sugerido = st.text_input("Nome do ficheiro Excel", value="Relatorio_Compras_Tapecaria")
        nome_final_xlsx = nome_sugerido if nome_sugerido.endswith(".xlsx") else f"{nome_sugerido}.xlsx"
            
    with st.expander("🏭 Fornecedores e Múltiplos"):
        st.caption("Edite ou adicione regras na última linha vazia.")
        df_regras_editado = st.data_editor(st.session_state.df_regras, num_rows="dynamic", use_container_width=True, hide_index=True)
        st.session_state.df_regras = df_regras_editado

# --- CORPO DO SITE ---
col1, col2 = st.columns([1, 15])
with col1:
    try: 
        st.image("simbolo.png", width=50)
    except: 
        pass
with col2:
    st.title("Inteligência de Compras")

st.markdown("##### Portal Operacional - Tapeçaria")
st.markdown("<br>", unsafe_allow_html=True)

# ====================================================
# === PROCESSAMENTO AUTOMÁTICO ("PISCOU, MUDOU") ===
# ====================================================
if uploaded_files:
    dfs_por_filial = {}
    todos_dados = []
    meses_globais = []
    dash_qtd_comprar = 0
    dash_qtd_transferida = 0
    dash_itens_pico = 0
    dash_itens_ruptura = 0
    
    for f in uploaded_files:
        df, meses = extrair_dados_pdf_web(f)
        if not df.empty:
            dfs_por_filial[f.name.replace(".pdf", "").upper()] = df
            todos_dados.append(df)
            if len(meses) >= 4 and not meses_globais: 
                meses_globais = meses[:4]
    
    if not meses_globais: 
        meses_globais = ["MÊS 1", "MÊS 2", "MÊS 3", "MÊS 4"]
    
    if not todos_dados:
        st.error("⚠️ O sistema não encontrou produtos compatíveis nos PDFs.")
    else:
        try:
            df_global = pd.concat(todos_dados).reset_index(drop=True)
            df_global['ESTOQUE_DISPONIVEL'] = df_global['ESTOQUE']
            
            vendas_recentes = df_global['MES_1'] + df_global['MES_2'] + df_global['MES_3'] + df_global['MES_4']
            df_global['TOTAL_VENDAS_RECENTES'] = vendas_recentes
            
            tracker_estoque = {}
            for _, row in df_global.iterrows():
                f_nome = row['FILIAL_NOME']
                c = row['CODIGO']
                est = float(row['ESTOQUE'])
                med = float(row['MEDIA_SISTEMA'])
                excesso = est if med == 0 else max(0.0, est - (med * meta))
                tracker_estoque[(f_nome, c)] = {'EXCEDENTE': excesso, 'MEDIA': med, 'ESTOQUE_FINAL': est}

            output = BytesIO()
            writer = pd.ExcelWriter(output, engine='openpyxl')
            
            for nome_destino, df_dest in dfs_por_filial.items():
                
                def classificar_estoque_parado(row):
                    if row['ESTOQUE'] > 0:
                        if row['MEDIA_SISTEMA'] == 0: return "🛑 SIM"
                        elif row['MESES_ESTOQUE'] > meses_parado: return "🛑 SIM"
                    return ""
                    
                ep_list = []
                for _, row in df_dest.iterrows():
                    ep_list.append(classificar_estoque_parado(row))
                df_dest['ESTOQUE PARADO'] = ep_list
                
                def processar_atipico(row):
                    meses_v = [row['MES_1'], row['MES_2'], row['MES_3'], row['MES_4']]
                    pico = max(meses_v)
                    outros = meses_v.copy()
                    outros.remove(pico)
                    soma_outros = sum(outros)
                    media_sem = soma_outros / 3 if soma_outros > 0 else 0
                    
                    media_sis = row['MEDIA_SISTEMA']
                    if media_sis > 0 and media_sem > 0: base_comp = min(media_sis, media_sem)
                    elif media_sem > 0: base_comp = media_sem
                    else: base_comp = media_sis
                        
                    if base_comp > 0 and pico >= (base_comp * fator_pico) and pico >= 30: 
                        return "⚠️ SIM", round(media_sem, 2)
                    elif base_comp == 0 and pico >= 30: 
                        return "⚠️ SIM", 0.0
                    return "Não", media_sis

                va_list = []
                mpc_list = []
                for _, row in df_dest.iterrows():
                    va, mpc = processar_atipico(row)
                    va_list.append(va)
                    mpc_list.append(mpc)
                df_dest['VENDA_ATIPICA'] = va_list
                df_dest['MEDIA_P_CALCULO'] = mpc_list
                
                dash_itens_pico += len(df_dest[df_dest['VENDA_ATIPICA'] == "⚠️ SIM"])
                
                # --- MELHORIA 3: DETEÇÃO DE RUPTURA CRÍTICA ---
                def classificar_ruptura(row):
                    if float(row['ESTOQUE']) == 0 and float(row['COMPRADA']) == 0 and float(row['MEDIA_P_CALCULO']) > 0:
                        return "🚨 CRÍTICA"
                    return "OK"
                
                rupt_list = []
                for _, row in df_dest.iterrows():
                    rupt_list.append(classificar_ruptura(row))
                df_dest['RUPTURA CRÍTICA'] = rupt_list
                dash_itens_ruptura += len(df_dest[df_dest['RUPTURA CRÍTICA'] == "🚨 CRÍTICA"])
                
                def calcular_log(row):
                    cod = row['CODIGO']
                    nec_calc = (row['MEDIA_P_CALCULO'] * meta) - (row['ESTOQUE'] + row['COMPRADA'])
                    necessidade = nec_calc
                    
                    if necessidade > 0:
                        opcoes = []
                        for f_outra in dfs_por_filial.keys():
                            if f_outra == nome_destino: continue
                            chave = (f_outra, cod)
                            if chave in tracker_estoque and tracker_estoque[chave]['EXCEDENTE'] > 0:
                                opcoes.append({'filial': f_outra, 'media': tracker_estoque[chave]['MEDIA'], 'excedente': tracker_estoque[chave]['EXCEDENTE']})
                        
                        if opcoes:
                            opcoes = sorted(opcoes, key=lambda x: (x['media'], -x['excedente']))
                            trans_item = []
                            nec_rest = necessidade
                            
                            for op in opcoes:
                                if nec_rest <= 0: break
                                chave_ced = (op['filial'], cod)
                                sal_ced = tracker_estoque[chave_ced]['EXCEDENTE']
                                
                                if sal_ced <= 0: continue
                                if nec_rest < 30 and sal_ced >= 30: qtd_a_tirar = 30
                                else: qtd_a_tirar = min(nec_rest, sal_ced)
                                    
                                if qtd_a_tirar >= 30:
                                    tracker_estoque[chave_ced]['EXCEDENTE'] -= qtd_a_tirar
                                    tracker_estoque[chave_ced]['ESTOQUE_FINAL'] -= qtd_a_tirar
                                    nec_rest -= qtd_a_tirar
                                    
                                    nome_ced = op['filial']
                                    if 'RIBEIR' in nome_ced: apelido = 'RP'
                                    elif 'LONDRINA' in nome_ced: apelido = 'Lon'
                                    elif 'FRANCA' in nome_ced: apelido = 'Frc'
                                    else: apelido = nome_ced
                                        
                                    trans_item.append(f"Tirar {int(qtd_a_tirar)} de {apelido}")
                                    
                            if trans_item: return " | ".join(trans_item), round(max(0, nec_rest), 2)
                    return "0", round(max(0, necessidade), 2)

                ti_list = []
                sug_list = []
                for _, row in df_dest.iterrows():
                    ti, sug = calcular_log(row)
                    ti_list.append(ti)
                    sug_list.append(sug)
                df_dest['TRANS INTERNA'] = ti_list
                sug_base = sug_list

                def aplicar_mult(row, sug):
                    if sug <= 0: return 0
                    forn = str(row.get('FORNECEDOR', '')).upper()
                    desc = str(row.get('DESCRICAO', '')).upper()
                    
                    for idx, regra in df_regras_editado.iterrows():
                        f_regra = str(regra.get('FORNECEDOR', '')).upper()
                        if f_regra and f_regra != "NAN" and f_regra in forn:
                            p_chave = str(regra.get('PALAVRA_CHAVE', '')).upper().strip()
                            if p_chave and p_chave != "NAN" and p_chave != "NONE":
                                if p_chave not in desc: continue 
                            try:
                                mult = int(regra['MULTIPLO'])
                                tol = int(regra['TOLERANCIA'])
                            except: continue 
                            
                            if mult > 0:
                                base = (int(sug) // mult) * mult
                                rest = sug % mult
                                if rest >= tol: return int(base + mult)
                                else: return int(base)
                    return int(math.ceil(sug))

                df_sug_zip = zip(df_dest.to_dict('records'), sug_base)
                df_dest['SUGESTAO COMPRA'] = [aplicar_mult(r, s) for r, s in df_sug_zip]
                dash_qtd_comprar += df_dest['SUGESTAO COMPRA'].sum()
                
                def extrair_n(t): 
                    if str(t) == "0": return 0
                    numeros = re.findall(r'\d+', str(t))
                    try: return sum([int(n) for n in numeros])
                    except: return 0
                    
                dash_qtd_transferida += df_dest['TRANS INTERNA'].apply(extrair_n).sum()
                
                renames = {'MES_1': meses_globais[0], 'MES_2': meses_globais[1], 'MES_3': meses_globais[2], 'MES_4': meses_globais[3], 'MEDIA_SISTEMA': 'MEDIA', 'MESES_ESTOQUE': 'MESES'}
                df_dest.rename(columns=renames, inplace=True)
                
                df_dest['ORIGINAL_SUGESTAO'] = df_dest['SUGESTAO COMPRA']
                df_dest['ORIGINAL_TRANS'] = df_dest['TRANS INTERNA']
                
                cols_f = [
                    'CODIGO', 'DESCRICAO', 'EMB.', meses_globais[0], meses_globais[1], meses_globais[2], meses_globais[3], 
                    'MEDIA', 'ESTOQUE', 'RESERVA', 'COMPRADA', 'MESES', 'SUGESTAO COMPRA', 'TRANS INTERNA', 
                    'VENDA_ATIPICA', 'ESTOQUE PARADO', 'RUPTURA CRÍTICA', 'ORIGINAL_SUGESTAO', 'ORIGINAL_TRANS'
                ]
                
                sheet_n = re.sub(r'[\\/*?:\[\]]', '', nome_destino)[:30]
                df_dest[cols_f].to_excel(writer, sheet_name=sheet_n, index=False)
                
                ws = writer.sheets[sheet_n]
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                header_font = Font(bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = 'A2'
                
                idx_compra = cols_f.index('SUGESTAO COMPRA') + 1
                idx_transf = cols_f.index('TRANS INTERNA') + 1
                idx_orig_sug = cols_f.index('ORIGINAL_SUGESTAO') + 1
                idx_orig_trans = cols_f.index('ORIGINAL_TRANS') + 1
                
                ws.column_dimensions[get_column_letter(idx_orig_sug)].hidden = True
                ws.column_dimensions[get_column_letter(idx_orig_trans)].hidden = True
                
                for col_idx, col in enumerate(ws.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in col:
                        cell.border = thin_border
                        if cell.row == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_align
                        elif col_idx != 2:
                            cell.alignment = center_align
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

                cv = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                ca = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
                cl = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
                cy = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                c_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                c_rup = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid") # Vermelho Ruptura
                
                idx_estoque = cols_f.index('ESTOQUE') + 1 
                idx_comprada = cols_f.index('COMPRADA') + 1 
                idx_atipica = cols_f.index('VENDA_ATIPICA') + 1
                idx_parado = cols_f.index('ESTOQUE PARADO') + 1
                idx_ruptura = cols_f.index('RUPTURA CRÍTICA') + 1
                
                for r in range(2, len(ws['A']) + 1):
                    if limpar_v(ws.cell(r, idx_comprada).value) > 0: ws.cell(r, idx_comprada).fill = cl 
                    if limpar_v(ws.cell(r, idx_compra).value) > 0: ws.cell(r, idx_compra).fill = cv 
                    if str(ws.cell(r, idx_transf).value) not in ["0", "None"]: ws.cell(r, idx_transf).fill = ca 
                    if "⚠️ SIM" in str(ws.cell(r, idx_atipica).value): ws.cell(r, idx_atipica).fill = cy 
                    
                    if "🛑 SIM" in str(ws.cell(r, idx_parado).value): 
                        ws.cell(r, idx_parado).fill = c_red
                        ws.cell(r, idx_estoque).fill = c_red
                        
                    if "🚨 CRÍTICA" in str(ws.cell(r, idx_ruptura).value):
                        ws.cell(r, idx_ruptura).fill = c_rup
                        ws.cell(r, idx_estoque).fill = c_rup

                max_row = len(ws['A'])
                if max_row >= 2:
                    yellow_cf_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    col_sug = get_column_letter(idx_compra)
                    col_orig_sug = get_column_letter(idx_orig_sug)
                    ws.conditional_formatting.add(f"{col_sug}2:{col_sug}{max_row}", FormulaRule(formula=[f"${col_sug}2<>${col_orig_sug}2"], stopIfTrue=False, fill=yellow_cf_fill))
                    
                    col_trans = get_column_letter(idx_transf)
                    col_orig_trans = get_column_letter(idx_orig_trans)
                    ws.conditional_formatting.add(f"{col_trans}2:{col_trans}{max_row}", FormulaRule(formula=[f"${col_trans}2<>${col_orig_trans}2"], stopIfTrue=False, fill=yellow_cf_fill))

            writer.close()

            def get_estoque_final(row):
                chave = (row['FILIAL_NOME'], row['CODIGO'])
                return tracker_estoque[chave]['ESTOQUE_FINAL'] if chave in tracker_estoque else row['ESTOQUE']
                
            edf_list = []
            for _, row in df_global.iterrows(): edf_list.append(get_estoque_final(row))
            df_global['ESTOQUE_DISPONIVEL'] = edf_list

            filtro_p1 = df_global['ESTOQUE_DISPONIVEL'] > 0
            filtro_p2 = df_global['MEDIA_SISTEMA'] == 0
            filtro_p3 = df_global['MESES_ESTOQUE'] > meses_parado
            df_p = df_global[filtro_p1 & (filtro_p2 | filtro_p3)]
            
            st.session_state.dfs_por_filial = dfs_por_filial
            st.session_state.dash_qtd_comprar = dash_qtd_comprar
            st.session_state.dash_qtd_transferida = dash_qtd_transferida
            st.session_state.dash_itens_pico = dash_itens_pico
            st.session_state.dash_itens_ruptura = dash_itens_ruptura
            st.session_state.df_p = df_p
            st.session_state.excel_data = output.getvalue()
            st.session_state.nome_final_xlsx = nome_final_xlsx
            st.session_state.analise_concluida = True

        except Exception as e:
            st.error(f"🚨 Ocorreu um erro interno durante os cálculos: {e}")
            st.code(traceback.format_exc())

    # --- RENDERIZAÇÃO DAS ABAS ---
    if st.session_state.analise_concluida:
        dfs_por_filial = st.session_state.dfs_por_filial
        dash_qtd_comprar = st.session_state.dash_qtd_comprar
        dash_qtd_transferida = st.session_state.dash_qtd_transferida
        dash_itens_pico = st.session_state.dash_itens_pico
        dash_itens_ruptura = st.session_state.dash_itens_ruptura
        df_p = st.session_state.df_p
        
        tab1, tab2, tab3, tab4 = st.tabs(["📊 Visão Geral", "🚨 Top Urgentes", "📦 Estoque Parado", "🔍 Prévia por Filial"])
        
        with tab1:
            st.subheader("Indicadores de Desempenho")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("🛒 Sugestão de Compra", f"{int(dash_qtd_comprar)} un.")
            c2.metric("🔄 Economia (Transf.)", f"{int(dash_qtd_transferida)} un.")
            c3.metric("⚠️ Picos de Vendas", f"{int(dash_itens_pico)} itens")
            c4.metric("🚨 Rupturas Críticas", f"{int(dash_itens_ruptura)} itens")
            
            if not df_p.empty: f_p = df_p.groupby('FILIAL_NOME')['ESTOQUE_DISPONIVEL'].sum().idxmax()
            else: f_p = "Nenhuma"
            c5.metric("📦 Maior Estoque Parado", f_p)
            
            st.success("✅ Inteligência processada com sucesso! O seu ficheiro Excel está pronto.")
            st.download_button(label="📥 Descarregar Relatório Inteligente (Excel)", data=st.session_state.excel_data, file_name=st.session_state.nome_final_xlsx, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        with tab2:
            df_all = pd.concat(dfs_por_filial.values())
            
            # BLOCO EXCLUSIVO DE RUPTURA CRÍTICA
            df_rupturas = df_all[df_all['RUPTURA CRÍTICA'] == "🚨 CRÍTICA"].sort_values(by='MEDIA', ascending=False)
            if not df_rupturas.empty:
                st.error("🚨 PRODUTOS EM RUPTURA CRÍTICA DETECTADOS (Estoque Zero + Sem Pedido em Andamento)")
                st.dataframe(df_rupturas[['CODIGO', 'DESCRICAO', 'FILIAL_NOME', 'MEDIA', 'SUGESTAO COMPRA', 'FORNECEDOR']], use_container_width=True)
            else:
                st.success("✅ Nenhuma ruptura crítica absoluta detectada nas filiais!")
                
            st.markdown("<br><hr>", unsafe_allow_html=True)
            st.subheader("🛒 Maior Volume de Compra Sugerido (Top 15)")
            top_compra = df_all[df_all['SUGESTAO COMPRA'] > 0].sort_values(by='SUGESTAO COMPRA', ascending=False).head(15)
            st.dataframe(top_compra[['CODIGO', 'DESCRICAO', 'FILIAL_NOME', 'SUGESTAO COMPRA', 'FORNECEDOR']], use_container_width=True)

        with tab3:
            st.subheader("Distribuição de Estoque Excedente")
            if not df_p.empty:
                grafico_dados = df_p.groupby('FILIAL_NOME')['ESTOQUE_DISPONIVEL'].sum().reset_index()
                fig = px.bar(grafico_dados, x='FILIAL_NOME', y='ESTOQUE_DISPONIVEL', title="Volume de Estoque Acima do Limite de Giro", color='ESTOQUE_DISPONIVEL', color_continuous_scale='Reds')
                st.plotly_chart(fig, use_container_width=True)
            else: st.info("Nenhum stock crítico detetado com base nos parâmetros atuais.")

        with tab4:
            st.subheader("Prévia Colorida dos Dados")
            sel_f = st.selectbox("Selecione a Filial para visualizar:", list(dfs_por_filial.keys()))
            df_view = dfs_por_filial[sel_f].copy()
            
            def pintar_tabela(row):
                cols = row.index
                estilos = [''] * len(cols)
                def f_idx(nome): return cols.get_loc(nome) if nome in cols else -1
                
                i_parado = f_idx('ESTOQUE PARADO')
                i_estoque = f_idx('ESTOQUE')
                i_atipica = f_idx('VENDA_ATIPICA')
                i_compra = f_idx('SUGESTAO COMPRA')
                i_transf = f_idx('TRANS INTERNA')
                i_comprada = f_idx('COMPRADA')
                i_ruptura = f_idx('RUPTURA CRÍTICA')
                
                if i_parado >= 0 and '🛑 SIM' in str(row.get('ESTOQUE PARADO', '')):
                    estilos[i_parado] = 'background-color: #F4CCCC; color: black;'
                    if i_estoque >= 0: estilos[i_estoque] = 'background-color: #F4CCCC; color: black;'
                if i_atipica >= 0 and '⚠️ SIM' in str(row.get('VENDA_ATIPICA', '')): estilos[i_atipica] = 'background-color: #FFF2CC; color: black;'
                if i_compra >= 0 and pd.to_numeric(row.get('SUGESTAO COMPRA', 0), errors='coerce') > 0: estilos[i_compra] = 'background-color: #D9EAD3; color: black;'
                if i_transf >= 0 and str(row.get('TRANS INTERNA', '')) not in ['0', 'None', '', 'nan']: estilos[i_transf] = 'background-color: #C9DAF8; color: black;'
                if i_comprada >= 0 and pd.to_numeric(row.get('COMPRADA', 0), errors='coerce') > 0: estilos[i_comprada] = 'background-color: #FCE5CD; color: black;'
                if i_ruptura >= 0 and '🚨 CRÍTICA' in str(row.get('RUPTURA CRÍTICA', '')):
                    estilos[i_ruptura] = 'background-color: #FFD2D2; color: black; font-weight: bold;'
                    if i_estoque >= 0: estilos[i_estoque] = 'background-color: #FFD2D2; color: black;'
                return estilos

            st.dataframe(df_view.style.apply(pintar_tabela, axis=1), use_container_width=True)

else: 
    st.info("A aguardar documentos. Por favor, carregue os ficheiros PDF na barra lateral para iniciar.")
